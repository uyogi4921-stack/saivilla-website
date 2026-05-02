import os
import io
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

logger = logging.getLogger(__name__)

SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
SMTP_USER = os.environ.get('SMTP_USER')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD')
NOTIFICATION_EMAIL = os.environ.get('NOTIFICATION_EMAIL')


def create_inquiry_excel(inquiry_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry Details"

    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    dark_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    label_font = Font(name="Arial", size=11, bold=True, color="D4AF37")
    value_font = Font(name="Arial", size=11, color="333333")
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "SAIVILLA DREAMHOUSE PVT.LTD."
    title_cell.font = title_font
    title_cell.fill = dark_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:B2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = "New Customer Inquiry"
    subtitle_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    subtitle_cell.fill = gold_fill
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    fields = [
        ("Customer Name", inquiry_data.get("name", "")),
        ("Email Address", inquiry_data.get("email", "")),
        ("Phone Number", inquiry_data.get("phone", "")),
        ("Property Interest", inquiry_data.get("propertyInterest") or "Not specified"),
        ("Message", inquiry_data.get("message", "")),
        ("Status", inquiry_data.get("status", "new").upper()),
        ("Inquiry Date", str(inquiry_data.get("createdAt", "N/A"))),
        ("Inquiry ID", inquiry_data.get("id", "")),
    ]

    for i, (label, value) in enumerate(fields, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = label_font
        label_cell.border = thin_border
        label_cell.alignment = Alignment(vertical="top")

        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.font = value_font
        value_cell.border = thin_border
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


async def send_inquiry_notification(inquiry_data: dict):
    try:
        message = MIMEMultipart('mixed')
        message['Subject'] = f'New Inquiry: {inquiry_data["propertyInterest"] or "General"}'
        message['From'] = SMTP_USER
        message['To'] = NOTIFICATION_EMAIL

        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #d4af37 0%, #f4d03f 100%);
                          color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                .content {{ background: #f9f9f9; padding: 20px; border: 1px solid #ddd; }}
                .field {{ margin: 15px 0; padding: 12px; background: white; border-radius: 4px;
                         border-left: 3px solid #d4af37; }}
                .label {{ font-weight: bold; color: #d4af37; margin-bottom: 5px; }}
                .value {{ color: #333; }}
                .footer {{ background: #333; color: white; padding: 15px; text-align: center;
                          border-radius: 0 0 8px 8px; font-size: 12px; }}
                .excel-note {{ background: #fff3cd; padding: 12px; border-radius: 4px;
                             margin-top: 15px; border: 1px solid #d4af37; font-size: 13px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2 style="margin: 0;">New Customer Inquiry</h2>
                    <p style="margin: 5px 0 0 0;">SAIVILLA DREAMHOUSE PVT.LTD.</p>
                </div>
                <div class="content">
                    <div class="field">
                        <div class="label">Customer Name</div>
                        <div class="value">{inquiry_data["name"]}</div>
                    </div>
                    <div class="field">
                        <div class="label">Email Address</div>
                        <div class="value">{inquiry_data["email"]}</div>
                    </div>
                    <div class="field">
                        <div class="label">Phone Number</div>
                        <div class="value">{inquiry_data["phone"]}</div>
                    </div>
                    <div class="field">
                        <div class="label">Property Interest</div>
                        <div class="value">{inquiry_data["propertyInterest"] or "Not specified"}</div>
                    </div>
                    <div class="field">
                        <div class="label">Message</div>
                        <div class="value">{inquiry_data["message"]}</div>
                    </div>
                    <div class="field">
                        <div class="label">Inquiry Time</div>
                        <div class="value">{inquiry_data.get("createdAt", "Just now")}</div>
                    </div>
                    <div class="excel-note">
                        <strong>Excel file attached</strong> - The inquiry details are also attached as an Excel spreadsheet for your records.
                    </div>
                </div>
                <div class="footer">
                    <p style="margin: 0;">SAIVILLA DREAMHOUSE PVT.LTD. - Palanpur, Gujarat</p>
                    <p style="margin: 5px 0 0 0;">+91 94263 19628 | saivillaoffice@gmail.com</p>
                </div>
            </div>
        </body>
        </html>
        """

        html_part = MIMEText(html_body, 'html')
        message.attach(html_part)

        excel_data = create_inquiry_excel(inquiry_data)
        customer_name = inquiry_data["name"].replace(" ", "_")
        filename = f"Inquiry_{customer_name}_{inquiry_data.get('id', 'unknown')[:8]}.xlsx"

        excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        excel_attachment.set_payload(excel_data)
        encoders.encode_base64(excel_attachment)
        excel_attachment.add_header('Content-Disposition', 'attachment', filename=filename)
        message.attach(excel_attachment)

        await aiosmtplib.send(
            message,
            hostname=SMTP_HOST,
            port=SMTP_PORT,
            username=SMTP_USER,
            password=SMTP_PASSWORD,
            start_tls=True
        )

        logger.info(f"Email notification with Excel sent successfully to {NOTIFICATION_EMAIL}")
        return True

    except Exception as e:
        logger.error(f"Failed to send email notification: {str(e)}")
        return False
